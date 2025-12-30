import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from django.db.models import Sum
from apps.sales.models import DailySale

def generate_sales_report_excel(start_date, end_date):
    """
    Gera relatório de vendas mensal em Excel.
    Retorna objeto BytesIO com o arquivo.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendas do Mês"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    money_format = 'R$ #,##0.00'
    
    # Cabeçalho
    headers = ["Data", "Vendedor", "Valor Total", "Observações"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Dados
    sales = DailySale.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).select_related('seller').order_by('date', 'seller__name')
    
    row_idx = 2
    for sale in sales:
        ws.cell(row=row_idx, column=1, value=sale.date).number_format = 'DD/MM/YYYY'
        ws.cell(row=row_idx, column=2, value=sale.seller.name)
        
        c_amount = ws.cell(row=row_idx, column=3, value=sale.amount)
        c_amount.number_format = money_format
        
        ws.cell(row=row_idx, column=4, value=sale.notes)
        row_idx += 1
        
    # Resumo por Vendedor
    ws_summary = wb.create_sheet("Resumo por Vendedor")
    ws_summary.append(["Vendedor", "Total Vendido", "Participação"])
    
    total_general = sales.aggregate(Sum('amount'))['amount__sum'] or 0
    
    ranking = (
        sales.values('seller__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )
    
    for item in ranking:
        percent = (item['total'] / total_general) if total_general else 0
        ws_summary.append([
            item['seller__name'],
            item['total'],
            f"{percent:.1%}"
        ])
    
    # Ajustar largura colunas
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column_letter].width = (max_length + 2) * 1.2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
